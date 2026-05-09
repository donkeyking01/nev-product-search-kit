import re
import requests
import json
import os
import time
import pandas as pd
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from colorama import Fore, init
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pathlib import Path

# 初始化 colorama（Windows 兼容）
init(autoreset=True)

def get_band_response(brand_id="0"):
    """获取品牌信息的响应"""
    num = 1
    while True:
        headers = {
            "user-agent": UserAgent().random
        }
        url = "https://car.autohome.com.cn/AsLeftMenu/As_LeftListNew.ashx"
        params = {
            "typeId": "1",
            "brandId": brand_id,
            "fctId": "0",
            "seriesId": "0"
        }
        response = requests.get(url, headers=headers, params=params)
        if response.status_code == 200:
            return response
        else:
            if num >= 5:
                print(Fore.RED + "请求超过5次，退出程序")
                return None
            else:
                print(Fore.YELLOW + "请求失败，正在重新请求...")
                num += 1
                time.sleep(1)

def get_response(series_id="0"):
    """获取车系配置信息的响应"""
    num = 1
    while True:
        headers = {
            "user-agent": UserAgent().random
        }
        url = "https://car-web-api.autohome.com.cn/car/param/getParamConf"
        params = {
            "mode": "1",
            "site": "1",
            "seriesid": series_id
        }
        response = requests.get(url, headers=headers, params=params)
        if response.status_code == 200:
            return response
        else:
            if num >= 5:
                print(Fore.RED + "请求超过5次，退出")
                return None
            else:
                print(Fore.YELLOW + "请求失败，正在重新请求...")
                num += 1
                time.sleep(1)

def get_car_config(config_dic):
    """从配置字典中提取车型配置数据"""
    allconfig = []
    configname_list = []
    for title in config_dic['result']['titlelist']:
        for item in title['items']:
            configname_list.append(item['itemname'])
    allconfig.append(configname_list)
    for data in config_dic['result']['datalist']:
        configvalue_list = []
        for valueitem in data['paramconflist']:
            if valueitem.get('itemname') != '':
                configvalue_list.append(valueitem['itemname'])
            elif not valueitem.get('sublist'):
                configvalue_list.append('-')
            else:
                stri = []
                for multivalue in valueitem['sublist']:
                    stri.append(multivalue['value'] + multivalue['name'])
                stro = '\n'.join(stri)
                configvalue_list.append(stro)
        allconfig.append(configvalue_list)
    return allconfig

def save_to_excel(data, folder, filename):
    """保存数据到 Excel 文件"""
    if not os.path.exists(folder):
        os.makedirs(folder)
    df = pd.DataFrame(data)
    excel_path = os.path.join(folder, filename)
    df.T.to_excel(excel_path, index=False, header=False)
    workbook = load_workbook(excel_path)
    sheet = workbook.active
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')
    num_columns = df.shape[0]
    for col in range(1, min(num_columns + 1, 702)):
        if col <= 26:
            sheet.column_dimensions[chr(64 + col)].width = 20
        else:
            first = chr(64 + (col - 1) // 26)
            second = chr(64 + (col - 1) % 26 + 1)
            sheet.column_dimensions[first + second].width = 20
    workbook.save(excel_path)
    print(Fore.GREEN + f"✓ 配置下载完成，保存到：{excel_path}\n")

def download_series_config(band, series_id, series_name):
    """下载指定车系的配置信息"""
    series_url = f"https://car.autohome.com.cn/config/series/{series_id}.html"
    print(Fore.CYAN + f"\n---正在下载 {band}-{series_name}，车型id为：{series_id}")
    print(Fore.CYAN + f"   配置链接：{series_url}")
    response = get_response(series_id)
    if not response:
        print(Fore.RED + f"✗ 获取 {series_name} 配置信息失败")
        return
    if "抱歉" in response.text and "暂无相关数据" in response.text:
        print(Fore.RED + f"✗ 该系列车暂无配置信息")
        return
    try:
        resp_dict = json.loads(response.text)
        all_info = get_car_config(resp_dict)
        output_dir = Path(__file__).parent.parent / "data" / "Raw" / "Para Raw" / band
        excel_name = f"{band}_{series_name}_配置.xlsx"
        save_to_excel(all_info, folder=str(output_dir), filename=excel_name)
    except Exception as e:
        print(Fore.RED + f"✗ 处理 {series_name} 配置数据时出错：{e}")

def parse_series(band, response):
    """解析品牌下的车系列表"""
    html = re.findall(r'document.writeln\("(.*)"\)', response.text)
    html = "".join(html)
    soup = BeautifulSoup(html, "html.parser")
    data_list = soup.select(".current > dl > dd > a")
    still_sell = [i for i in data_list if "停售" not in i.get_text(strip=True)]
    stop_sell = [i for i in data_list if "停售" in i.get_text(strip=True)]
    print(Fore.CYAN + f"\n该品牌共找到{len(data_list)}个车型，其中：")
    print(Fore.CYAN + f"  在售车型：{len(still_sell)} 个")
    print(Fore.CYAN + f"  停售车型：{len(stop_sell)} 个（停售车型无配置信息）")
    print(Fore.CYAN + "-"*60)
    print(Fore.CYAN + "在售车型列表：")
    print(Fore.CYAN + "-"*60)
    series_dict = {}
    for still_index, still_data in enumerate(still_sell, start=1):
        series_name = still_data.contents[0].text.strip()
        href = still_data.get("href")
        series_id = re.findall(r'/price/series-(\d+).html', href)[0]
        series_dict[series_id] = series_name
        print(f"序号：{still_index}\t车型：{series_name}\t车型id：{series_id}")
    while True:
        choice = input(Fore.YELLOW + "\n请输入需要下载的车型id，输入0则下载该品牌全部车型配置：").strip()
        if choice == "0":
            print(Fore.CYAN + f"\n开始下载 {band} 的所有车型配置...\n")
            for series_id, series_name in series_dict.items():
                download_series_config(band, series_id, series_name)
                time.sleep(2)
            break
        elif choice in series_dict.keys():
            series_name = series_dict[choice]
            download_series_config(band, choice, series_name)
            break
        else:
            print(Fore.RED + "输入的车型id不存在，请重新输入。")
            continue
    input(Fore.GREEN + "\n下载完成！请按回车键关闭程序...")

def main():
    """主函数 - 交互式获取品牌和车型信息"""
    print(Fore.CYAN + "\n" + "="*60)
    print(Fore.CYAN + "        汽车之家车型参数配置爬虫")
    print(Fore.CYAN + "="*60 + "\n")
    while True:
        band = input(Fore.YELLOW + "请输入汽车品牌（如：比亚迪、问界）：").strip()
        if not band:
            print(Fore.RED + "品牌名称不能为空，请重新输入")
            continue
        print(Fore.CYAN + "\n正在查询品牌信息...")
        response = get_band_response()
        if not response:
            print(Fore.RED + "获取品牌列表失败，请检查网络连接")
            continue
        band_pattern = f"<a href=([^>]*?)><i[^>]*?></i>{band}<em>"
        band_info = re.search(band_pattern, response.text)
        if not band_info:
            print(Fore.RED + f"未找到品牌 '{band}'，请检查拼写或重新输入")
            continue
        else:
            band_href = band_info.group(1)
            band_id = re.findall(r'/price/brand-(\d+).html', band_href)[0]
            print(Fore.GREEN + f"✓ 找到品牌：{band}，品牌ID：{band_id}")
            print(Fore.CYAN + "正在获取车型列表...")
            resp_brand = get_band_response(brand_id=band_id)
            if not resp_brand:
                print(Fore.RED + "获取品牌详情失败")
                continue
            parse_series(band, response=resp_brand)
            break

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print(Fore.YELLOW + "\n\n程序已被用户中断")
    except Exception as e:
        print(Fore.RED + f"\n程序运行出错：{e}")
        input("按回车键退出...")
